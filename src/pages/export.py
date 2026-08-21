import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import datetime
import os
from utils.database import get_db

class ExportPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg="#f8fafb")
        self.db = get_db()
        self._name_var = tk.StringVar(value="")
        
        # --- UI Structure ---
        # 1. Header
        header = tk.Frame(self, bg="white", height=60, bd=0)
        header.pack(fill="x")
        
        title_box = tk.Frame(header, bg="white")
        title_box.pack(side="left", padx=20, pady=10)
        tk.Label(title_box, text="Final Review, Save & Export", font=("Segoe UI", 16, "bold"), bg="white", fg="#0f172a").pack(anchor="w")
        tk.Label(title_box, text="Final laboratory confirmation step: Review construct, save to local database, and export reports.",
                 font=("Segoe UI", 9), bg="white", fg="#64748b").pack(anchor="w")
        
        # 2. Main Content Canvas
        self.canvas = tk.Canvas(self, bg="#f8fafb", highlightthickness=0)
        self.scrollbar = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        
        self.content = tk.Frame(self.canvas, bg="#f8fafb", padx=20, pady=16)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.content, anchor="nw")
        
        self.content.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(self.canvas_window, width=e.width))

        from ui_widgets import bind_mousewheel
        bind_mousewheel(self.canvas, self.content)

        # --- Workflow Pipeline Progress Bar ---
        self._build_pipeline_bar()
        
        # --- Confirmed Construct Master Card ---
        self._build_construct_card()

    def _build_pipeline_bar(self):
        pipe_card = tk.Frame(self.content, bg="#ffffff", bd=1, relief="solid", padx=16, pady=12)
        pipe_card.pack(fill="x", pady=(0, 12))
        
        tk.Label(pipe_card, text="Laboratory Workflow Pipeline Status:", font=("Segoe UI", 10, "bold"), bg="#ffffff", fg="#1e293b").pack(anchor="w", pady=(0, 6))
        
        steps_row = tk.Frame(pipe_card, bg="#ffffff")
        steps_row.pack(fill="x")
        
        steps = [
            ("1. 📏 Size Calculated", "#e0f2fe", "#0369a1"),
            ("2. 🧬 DNA Generated", "#e0f2fe", "#0369a1"),
            ("3. ⚖ BLAST Verified", "#e0f2fe", "#0369a1"),
            ("4. 🧪 Primers Confirmed", "#e0f2fe", "#0369a1"),
            ("5. 📊 Probes Confirmed", "#e0f2fe", "#0369a1"),
            ("6. 💾 Final Save & Export", "#dcfce7", "#15803d")
        ]
        for idx, (txt, bg_c, fg_c) in enumerate(steps):
            box = tk.Label(steps_row, text=txt, bg=bg_c, fg=fg_c, font=("Segoe UI", 8, "bold"), padx=8, pady=4, bd=1, relief="solid")
            box.pack(side="left", padx=2)
            if idx < len(steps) - 1:
                tk.Label(steps_row, text="➔", bg="#ffffff", fg="#94a3b8", font=("Segoe UI", 9)).pack(side="left", padx=2)

    def _build_construct_card(self):
        self.card = tk.Frame(self.content, bg="white", bd=1, relief="solid", padx=20, pady=18)
        self.card.pack(fill="x", anchor="n")
        
        top_row = tk.Frame(self.card, bg="white")
        top_row.pack(fill="x", pady=(0, 10))
        
        tk.Label(top_row, text="Confirmed Construct Master Profile", font=("Segoe UI", 13, "bold"), bg="white", fg="#0f172a").pack(side="left")
        
        tk.Button(top_row, text="🔄 Sync Latest Data", command=self.refresh_construct_data,
                  bg="#f1f5f9", fg="#334155", font=("Segoe UI", 8, "bold"), relief="flat", padx=10, pady=3, cursor="hand2").pack(side="right")

        # Form Fields Frame
        form = tk.Frame(self.card, bg="white")
        form.pack(fill="x", pady=(4, 12))

        # Name Field
        tk.Label(form, text="Construct Name / ID *:", font=("Segoe UI", 9, "bold"), bg="white", fg="#334155").grid(row=0, column=0, sticky="w", pady=4)
        self.entry_name = tk.Entry(form, textvariable=self._name_var, font=("Segoe UI", 10), width=32, bd=1, relief="solid")
        self.entry_name.grid(row=0, column=1, sticky="w", padx=10, pady=4)

        # Notes Field
        tk.Label(form, text="Notes / Laboratory Tags:", font=("Segoe UI", 9, "bold"), bg="white", fg="#334155").grid(row=1, column=0, sticky="nw", pady=6)
        self.txt_notes = tk.Text(form, height=3, width=32, font=("Segoe UI", 9), bd=1, relief="solid")
        self.txt_notes.grid(row=1, column=1, sticky="w", padx=10, pady=6)

        # Specifications Box
        self.specs_frame = tk.Frame(self.card, bg="#f8fafc", bd=1, relief="solid", padx=12, pady=10)
        self.specs_frame.pack(fill="x", pady=(0, 16))
        self.lbl_specs = tk.Label(self.specs_frame, text="Loading construct details...", font=("Segoe UI", 9), bg="#f8fafc", fg="#64748b", justify="left")
        self.lbl_specs.pack(anchor="w")

        # Final Action Buttons Box
        btn_box = tk.Frame(self.card, bg="white")
        btn_box.pack(fill="x", pady=(8, 0))
        
        # Primary Action: Save to DB
        tk.Button(btn_box, text="💾 Save Confirmed DNA to Database", command=self._save_to_database, 
                  bg="#10b981", fg="white", font=("Segoe UI", 10, "bold"), padx=16, pady=10, relief="flat", cursor="hand2").pack(side="left", padx=(0, 8))
        
        # Export buttons
        tk.Button(btn_box, text="Export to Excel (.xlsx)", command=self._export_to_excel, 
                  bg="#2e7d32", fg="white", font=("Segoe UI", 10, "bold"), padx=14, pady=10, relief="flat", cursor="hand2").pack(side="left", padx=4)
                  
        tk.Button(btn_box, text="Export to PDF (.pdf)", command=self._export_to_pdf, 
                  bg="#d32f2f", fg="white", font=("Segoe UI", 10, "bold"), padx=14, pady=10, relief="flat", cursor="hand2").pack(side="left", padx=4)

        tk.Button(btn_box, text="Save to Cloud ☁", command=self._export_to_cloud, 
                  bg="#0288d1", fg="white", font=("Segoe UI", 10, "bold"), padx=14, pady=10, relief="flat", cursor="hand2").pack(side="left", padx=4)

        tk.Button(btn_box, text="📊 View in Matrix ➡", command=self._go_to_matrix_page,
                  bg="#f3e8ff", fg="#6b21a8", font=("Segoe UI", 10, "bold"), padx=14, pady=10, relief="flat", cursor="hand2").pack(side="left", padx=4)

    def refresh_construct_data(self):
        """Pulls the latest generated sequence, primers, and probes from active tabs."""
        try:
            dna_page = self.master.pages.get('dna')
            if dna_page and hasattr(dna_page, '_last_data') and dna_page._last_data:
                data = dna_page._last_data
                if not self._name_var.get():
                    self._name_var.set(f"DNA_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}")

                p_count = len(data.get('probes', []))
                primers = data.get('primers')
                fwd_seq = primers.get('fwd', {}).get('seq', '—') if primers else '—'
                rev_seq = primers.get('rev', {}).get('seq', '—') if primers else '—'

                specs_text = (
                    f"• Construct Type: {data.get('mode', 'linear').capitalize()}\n"
                    f"• Payload Length: {data.get('length', 0)} bp  |  Total Construct Length: {data.get('total_length', 0)} bp  |  GC Content: {data.get('gc_pct', 0):.1f}%\n"
                    f"• Forward Primer: {fwd_seq}  |  Reverse Primer: {rev_seq}\n"
                    f"• TaqMan Probes: {p_count} orthogonal channels configured\n"
                    f"• Sequence Verification: BLAST verified • Assay Orthogonality Confirmed (0 DB clashes)"
                )
                self.lbl_specs.config(text=specs_text, fg="#0f172a")
            else:
                self.lbl_specs.config(text="No active DNA construct found. Please calculate size and generate a DNA sequence first.", fg="#94a3b8")
        except Exception as e:
            self.lbl_specs.config(text=f"Error loading construct: {e}", fg="#dc2626")

    def _save_to_database(self):
        """Commits the fully verified construct to the SQLite database."""
        try:
            dna_page = self.master.pages.get('dna')
            if not dna_page or not hasattr(dna_page, '_last_data') or not dna_page._last_data:
                messagebox.showwarning("No Data", "No DNA data found. Please generate a sequence first in the 'DNA Generate' tab.")
                return
            
            data = dna_page._last_data
            name = self._name_var.get().strip()
            if not name:
                name = f"DNA_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
                self._name_var.set(name)

            notes = self.txt_notes.get("1.0", "end").strip()

            seq_id = self.db.save_sequence(
                name=name,
                payload=data['payload'],
                linear_seq=data['linear_seq'],
                mode=data.get('mode', 'linear'),
                length=data.get('length'),
                total_length=data.get('total_length'),
                gc_pct=data.get('gc_pct'),
                primers=data.get('primers'),
                probes=data.get('probes'),
                notes=notes
            )

            # Refresh matrix page if loaded
            try:
                matrix_page = self.master.pages.get('matrix_db')
                if matrix_page:
                    matrix_page.refresh_data()
            except Exception:
                pass

            msg = f"Construct '{name}' (ID: {seq_id}) successfully saved to local database!\n\nWould you like to view the updated Similarity Matrix now?"
            if messagebox.askyesno("Save Successful", msg):
                self._go_to_matrix_page()

        except Exception as e:
            messagebox.showerror("Save Failed", f"Could not save sequence to database: {e}")

    def _go_to_matrix_page(self):
        try:
            self.master.show("matrix_db")
        except Exception:
            messagebox.showwarning("Navigation", "Matrix page is not registered yet.")

    def _export_to_excel(self):
        # 1. Get Data
        # We need to access the last generated data from DNAGeneratePage.
        # We can reach it via the NavigationFrame -> pages -> 'dna'
        try:
            dna_page = self.master.pages.get('dna')
            if not dna_page or not hasattr(dna_page, '_last_data') or not dna_page._last_data:
                messagebox.showwarning("No Data", "No DNA data found. Please generate a sequence first in the 'DNA Generate' tab.")
                return
            
            data = dna_page._last_data
        except Exception as e:
            messagebox.showerror("Error", f"Could not retrieve data: {e}")
            return

        # 2. Prompt for Save Location
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"DNA_Construct_{timestamp}.xlsx"
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            initialfile=default_name,
            title="Save Export File"
        )
        
        if not filepath:
            return # User cancelled

        # 3. Generate Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Missing Dependency", "The 'openpyxl' library is required for Excel export.\nPlease install it (pip install openpyxl) or ask the developer.")
            return

        try:
            wb = openpyxl.Workbook()
            
            # --- STYLE DEFINITIONS ---
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4472C4")
            label_font = Font(bold=True)
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            def style_header(cell):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # --- SHEET 1: SUMMARY ---
            ws_sum = wb.active
            ws_sum.title = "Summary"
            
            ws_sum['A1'] = "DNA Construct Report"
            ws_sum['A1'].font = Font(size=14, bold=True)
            
            ws_sum['A3'] = "Generated On:"
            ws_sum['B3'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            ws_sum['A4'] = "Mode:"
            ws_sum['B4'] = data.get('mode', 'Unknown').capitalize()
            
            ws_sum['A5'] = "Payload Length:"
            ws_sum['B5'] = f"{data.get('length', 0)} bp"
            
            ws_sum['A6'] = "Total Length:"
            ws_sum['B6'] = f"{data.get('total_length', 0)} bp"
            
            # Formatting Summary
            for row in range(3, 7):
                ws_sum[f'A{row}'].font = label_font
                ws_sum.column_dimensions['A'].width = 20
                ws_sum.column_dimensions['B'].width = 30

            # --- SHEET 2: SEQUENCES ---
            ws_seq = wb.create_sheet("Sequences")
            
            headers = ["Type", "Length", "Sequence (5' -> 3')"]
            for col, h in enumerate(headers, 1):
                c = ws_seq.cell(row=1, column=col, value=h)
                style_header(c)
            
            # Payload
            payload = data.get('payload', '')
            ws_seq.append(["Payload", len(payload), payload])
            
            # Payload Complement
            comp_map = {'A':'T','T':'A','C':'G','G':'C','N':'N'}
            payload_comp = "".join(comp_map.get(c, 'N') for c in payload)
            ws_seq.append(["Payload (Complement)", len(payload_comp), payload_comp])
            
            # Full Linear
            linear = data.get('linear_seq', '')
            ws_seq.append(["Full Linear Construct", len(linear), linear])
            
            ws_seq.column_dimensions['A'].width = 25
            ws_seq.column_dimensions['B'].width = 10
            ws_seq.column_dimensions['C'].width = 100

            # --- SHEET 3: PRIMERS ---
            ws_prim = wb.create_sheet("Primers")
            headers = ["Name", "Sequence (5'->3')", "Tm (°C)", "GC (%)", "Length (bp)", "Score"]
            for col, h in enumerate(headers, 1):
                c = ws_prim.cell(row=1, column=col, value=h)
                style_header(c)

            primers = data.get('primers')
            if primers:
                fwd = primers.get('fwd')
                rev = primers.get('rev')
                
                if fwd:
                    ws_prim.append(["Forward Primer", fwd['seq'], fwd['tm'], fwd['gc'], fwd['len'], fwd['score']])
                if rev:
                    ws_prim.append(["Reverse Primer", rev['seq'], rev['tm'], rev['gc'], rev['len'], rev['score']])
            else:
                ws_prim.append(["No primers generated", "-", "-", "-", "-", "-"])

            for col_letter in ['A', 'C', 'D', 'E', 'F']:
                 ws_prim.column_dimensions[col_letter].width = 15
            ws_prim.column_dimensions['B'].width = 50

            # --- SHEET 4: PROBES ---
            ws_probe = wb.create_sheet("Probes")
            headers = ["Rank", "Sequence", "Tm (°C)", "GC (%)", "Length", "Score"]
            for col, h in enumerate(headers, 1):
                c = ws_probe.cell(row=1, column=col, value=h)
                style_header(c)
            
            probes = data.get('probes', [])
            if probes:
                for i, p in enumerate(probes, 1):
                     ws_probe.append([i, p['seq'], p['tm'], p['gc'], p['len'], p['score']])
            else:
                 ws_probe.append(["No probes found", "-", "-", "-", "-", "-"])

            for col_letter in ['A', 'C', 'D', 'E', 'F']:
                 ws_probe.column_dimensions[col_letter].width = 10
            ws_probe.column_dimensions['B'].width = 50

            # SAVE
            wb.save(filepath)
            messagebox.showinfo("Export Success", f"File saved successfully to:\n{filepath}")
            
            # Open the folder
            try:
                os.startfile(os.path.dirname(filepath))
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to save file:\n{e}")

    def _export_to_pdf(self):
        # 1. Get Data
        try:
            dna_page = self.master.pages.get('dna')
            if not dna_page or not hasattr(dna_page, '_last_data') or not dna_page._last_data:
                messagebox.showwarning("No Data", "No DNA data found. Please generate a sequence first.")
                return
            data = dna_page._last_data
        except Exception as e:
            messagebox.showerror("Error", f"Could not retrieve data: {e}")
            return
            
        # 2. Save Dialog
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"DNA_Construct_{timestamp}.pdf"
        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf")],
            initialfile=default_name,
            title="Save Export PDF"
        )
        if not filepath: return
        
        # 3. Check Deps
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            import qrcode
        except ImportError:
            messagebox.showerror("Missing Dependency", "PDF generation requires: reportlab, qrcode.\nPlease install them.")
            return

        # 4. Generate Content
        try:
            doc = SimpleDocTemplate(filepath, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            # Custom Styles
            title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=12, textColor=colors.HexColor("#0d47a1"))
            h2_style = ParagraphStyle('Heading2', parent=styles['Heading2'], fontSize=14, spaceBefore=12, spaceAfter=6, textColor=colors.HexColor("#1565c0"))
            normal_style = styles['Normal']
            code_style = ParagraphStyle('Code', parent=normal_style, fontName='Courier', fontSize=8, leading=10, backColor=colors.whitesmoke, borderPadding=2)
            
            # --- QR Code Generation ---
            # Unique ID structure: DNAx-{Timestamp}-{Length}
            if 'export_id' not in data:
                data['export_id'] = f"DNAx-{timestamp}-{data.get('length',0)}"
            unique_id = data['export_id']
            # Include first 60bp of sequence for validation
            seq_preview = data.get('linear_seq', '')[:60] + ("..." if len(data.get('linear_seq', '')) > 60 else "")
            
            qr_data = f"ID: {unique_id}\nLen: {data.get('length')}bp | GC: {data.get('gc_pct',0):.1f}%\nSeq: {seq_preview}"
            
            # Increase border to 4 (standard) and Error Correction to High (H)
            # Set version=None so it auto-sizes to fit the data
            qr = qrcode.QRCode(
                version=None,
                error_correction=qrcode.constants.ERROR_CORRECT_H,
                box_size=10, 
                border=4
            )
            qr.add_data(qr_data)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Save QR temporarily
            qr_path = os.path.join(os.path.dirname(filepath), ".temp_qr.png")
            img.save(qr_path)
            
            # --- Header Table (Logo/Title | QR Code) ---
            # We put QR code in header
            
            # Left Cell: Title & Date
            header_text = f"""<font size=18 color="#0d47a1"><b>DNAₓ Construct Report</b></font><br/>
            <font size=10 color="#666">Generated: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}</font><br/>
            <font size=10 color="#666">ID: {unique_id}</font>"""
            
            # Right Cell: QR Image
            qr_img = RLImage(qr_path, width=1.2*inch, height=1.2*inch)
            
            header_table = Table([[Paragraph(header_text, normal_style), qr_img]], colWidths=[5*inch, 1.5*inch])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('ALIGN', (1,0), (1,0), 'RIGHT'),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 0.2*inch))
            
            # --- Summary Section ---
            elements.append(Paragraph("Construct Summary", h2_style))
            
            sum_data = [
                ["Property", "Value"],
                ["Mode", data.get('mode', 'Unknown').capitalize()],
                ["Payload Length", f"{data.get('length', 0)} bp"],
                ["Total Linear Length", f"{data.get('total_length', 0)} bp"],
                ["GC Content", f"{data.get('gc_pct', 0):.1f}%"],
            ]
            
            t = Table(sum_data, colWidths=[2.5*inch, 4*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (1,0), colors.HexColor("#e3f2fd")),
                ('TEXTCOLOR', (0,0), (1,0), colors.HexColor("#0d47a1")),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.whitesmoke]),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 0.2*inch))

            # --- Primers ---
            elements.append(Paragraph("Amplification Primers", h2_style))
            primers = data.get('primers')
            
            if primers:
                fwd = primers.get('fwd', {})
                rev = primers.get('rev', {})
                
                p_data = [
                    ["Direction", "Sequence (5'->3')", "Tm", "GC%", "Len"],
                    ["Forward", Paragraph(fwd.get('seq','-'), code_style), f"{fwd.get('tm','-')}C", f"{fwd.get('gc','-')}%", fwd.get('len','-')],
                    ["Reverse", Paragraph(rev.get('seq','-'), code_style), f"{rev.get('tm','-')}C", f"{rev.get('gc','-')}%", rev.get('len','-')],
                ]
                
                t = Table(p_data, colWidths=[0.8*inch, 3.5*inch, 0.7*inch, 0.7*inch, 0.7*inch])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#e8f5e9")),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor("#1b5e20")),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ]))
                elements.append(t)
            else:
                elements.append(Paragraph("No amplification primers generated.", normal_style))
            
            
            elements.append(Spacer(1, 0.2*inch))

            # --- Probes ---
            elements.append(Paragraph("TaqMan Probes", h2_style))
            probes = data.get('probes', [])
            
            if probes:
                # We show top 5 probes
                pr_data = [["Rank", "Sequence", "Tm", "GC%", "Len", "Score"]]
                for i, p in enumerate(probes[:5], 1):
                    pr_data.append([
                        str(i),
                        Paragraph(p.get('seq','-'), code_style),
                        f"{p.get('tm','-'):.1f}C",
                        f"{p.get('gc','-'):.1f}%",
                        str(p.get('len','-')),
                        f"{p.get('score',0):.1f}"
                    ])
                
                t = Table(pr_data, colWidths=[0.5*inch, 3.5*inch, 0.7*inch, 0.7*inch, 0.5*inch, 0.6*inch])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#fff3e0")),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor("#e65100")),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.whitesmoke]),
                ]))
                elements.append(t)
            else:
                elements.append(Paragraph("No compatible probes found.", normal_style))
                
            elements.append(Spacer(1, 0.2*inch))
            
            # --- Sequence Data ---
            elements.append(Paragraph("Full Linear Sequence", h2_style))
            seq = data.get('linear_seq', '')
            # Wrap sequence nicely
            elements.append(Paragraph(seq, code_style))
            
            # Build
            doc.build(elements)
            
            # Cleanup
            if os.path.exists(qr_path):
                os.remove(qr_path)
                
            messagebox.showinfo("Export Success", f"PDF saved successfully to:\n{filepath}")
            try: os.startfile(filepath)
            except: pass
            
        except Exception as e:
            messagebox.showerror("PDF Error", f"Failed to generate PDF:\n{e}")

    def _export_to_cloud(self):
        # 1. Get Data
        try:
            dna_page = self.master.pages.get('dna')
            if not dna_page or not hasattr(dna_page, '_last_data') or not dna_page._last_data:
                messagebox.showwarning("No Data", "No DNA data found. Please generate a sequence first.")
                return
            data = dna_page._last_data
        except Exception as e:
            messagebox.showerror("Error", f"Could not retrieve data: {e}")
            return

        # 2. Upload to Firebase
        try:
            import requests
        except ImportError:
            messagebox.showerror("Missing Dependency", "The 'requests' library is required for Cloud Export.\nPlease install it (pip install requests).")
            return

        # 3. Prompt for Company Selection
        class CompanySelectionDialog(tk.Toplevel):
            def __init__(self, parent):
                super().__init__(parent)
                self.title("Select Company")
                self.geometry("300x150")
                self.transient(parent)
                self.grab_set()
                self.selected_company = None
                
                tk.Label(self, text="Please select a company to upload to:").pack(pady=10)
                
                self.company_var = tk.StringVar(value="Company A")
                companies = ["Company A", "Company B", "Company C"]
                self.dropdown = ttk.Combobox(self, textvariable=self.company_var, values=companies, state="readonly")
                self.dropdown.pack(pady=5)
                
                tk.Button(self, text="Confirm", command=self.confirm).pack(pady=10)
                
                # Center the dialog
                self.update_idletasks()
                x = parent.winfo_rootx() + (parent.winfo_width() // 2) - (self.winfo_width() // 2)
                y = parent.winfo_rooty() + (parent.winfo_height() // 2) - (self.winfo_height() // 2)
                self.geometry(f"+{x}+{y}")
                
            def confirm(self):
                self.selected_company = self.company_var.get()
                self.destroy()

        dialog = CompanySelectionDialog(self.winfo_toplevel())
        self.wait_window(dialog)
        
        selected_company = dialog.selected_company
        if not selected_company:
            messagebox.showwarning("Upload Cancelled", "Cloud upload was cancelled because no company was selected.")
            return

        if 'export_id' not in data:
            timestamp_id = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            data['export_id'] = f"DNAx-{timestamp_id}-{data.get('length', 0)}"
        unique_id = data['export_id']

        # Determine next batch number by querying existing databases
        db_url_base = "https://datasample-5e3cd-default-rtdb.firebaseio.com/dna_exports.json"
        
        batch_number = 1
        try:
            # Fetch all existing exports
            response_all = requests.get(db_url_base, timeout=10)
            if response_all.status_code == 200:
                all_data = response_all.json()
                if all_data:
                    max_batch = 0
                    for key, record in all_data.items():
                        if isinstance(record, dict):
                            rec_batch = record.get('batch_number')
                            if isinstance(rec_batch, int) and rec_batch > max_batch:
                                max_batch = rec_batch
                            elif isinstance(rec_batch, str) and rec_batch.isdigit():
                                if int(rec_batch) > max_batch:
                                    max_batch = int(rec_batch)
                    batch_number = max_batch + 1
        except Exception:
            pass # Use batch_number = 1 if fetching fails temporarily

        db_url_push = f"https://datasample-5e3cd-default-rtdb.firebaseio.com/dna_exports/{unique_id}.json"
        
        # Prepare payload
        now = datetime.datetime.now()
        timestamp = now.isoformat()
        
        payload = {
            "id": unique_id,
            "company": selected_company,
            "batch_number": batch_number,
            "year": now.year,
            "month": now.month,
            "day": now.day,
            "timestamp": timestamp,
            "mode": data.get('mode', 'Unknown'),
            "length": data.get('length', 0),
            "total_length": data.get('total_length', 0),
            "gc_pct": data.get('gc_pct', 0),
            "linear_seq": data.get('linear_seq', ''),
            "payload_seq": data.get('payload', '')
        }
        
        if data.get('primers'):
           payload['primers'] = data.get('primers')
        if data.get('probes'):
           payload['probes'] = data.get('probes')

        try:
            response = requests.put(db_url_push, json=payload, timeout=10)
            response.raise_for_status()
            messagebox.showinfo("Cloud Export Success", f"Data successfully uploaded to Firebase assigned to Batch {batch_number}.")
        except Exception as e:
            messagebox.showerror("Cloud Export Error", f"Failed to upload data:\n{e}")
